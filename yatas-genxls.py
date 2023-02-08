import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

results = yaml.safe_load(open("results.yaml", 'r'))

checks_info = yaml.safe_load(open("checks.yaml", 'r'))['checks']

wb = Workbook()
ws = wb.active

ws.append(["Category", "Reference", "Macro", "Priority", "Time", "Complexity", "Resources"])
for i in range(ws.max_column):
    ws.cell(1,i+1).font = Font(bold=True)
    ws.cell(1,i+1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(1,i+1).border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

ws.append([
    "Name of the category",
    "",
    "Subtasks",
    """Priority according to the security risks
- Low : Non-priority tasks that do not address a risk but are good practices
- Medium : Non-priority tasks that address a minor risk
- High : Tasks to be scheduled quickly that involve a risk of partial compromised
- Critical : Tasks to be scheduled quickly that involve a risk of total compromised""",
    """Estimated implementation
- S :  <3 days
- M : < 1 week
- L : < 1,5 week
- XL : > 1,5 week
These are approximate estimates""",
    "Scale from 1 to 4 to evaluate the difficulty of the implementation",
    ""
])

ws.row_dimensions[2].height = 100
for i in range(ws.max_column):
    ws.cell(2,i+1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws.cell(2,i+1).border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='medium', color='000000'))
ws.cell(2,4).alignment = Alignment(horizontal='left', vertical='center', wrapText=False)
ws.cell(2,5).alignment = Alignment(horizontal='left', vertical='center', wrapText=False)

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 100

for check in results[0]['checks']:
    if check['status'] != "OK":
        failed_resources = [res['resource_arn'] for res in check['results'] if res['status'] != "OK"]
        ws.append([
            checks_info[check['id']]['category'],
            check['id'],
            checks_info[check['id']]['reco'],
            "",
            "",
            "",
            "\n".join(failed_resources if len(failed_resources) <= 6 else failed_resources[:3]+[f"... and {len(failed_resources)-3} more"])
        ])
        if len(failed_resources) > 6:
            new_ws = wb.create_sheet(check['id'])
            new_ws.append([checks_info[check['id']]['reco']])
            for res in failed_resources:
                new_ws.append([res])
        for i in range(ws.max_column):
            ws.cell(ws.max_row,i+1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(ws.max_row,i+1).border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
        

wb.save("results.xlsx")
